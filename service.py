import openpyxl
import sys
from config import template_file
from excel_processing import process_sheet, create_styled_homepage

# --- MAIN SCRIPT ---
def excel_processing(raw_file, output_file, SHEETS_TO_IGNORE, LOGO_FILENAME):
    print("Starting Data Transfer")
    try:
        raw_wb = openpyxl.load_workbook(raw_file, data_only=True)
        template_wb = openpyxl.load_workbook(template_file)
    except FileNotFoundError as e:
        print(f"ERROR: Could not find a required file: {e.filename}")
        sys.exit()

    process_sheet(raw_wb, template_wb, SHEETS_TO_IGNORE)

    print("Processing and writing data one sheet at a time")
    create_styled_homepage(template_wb, SHEETS_TO_IGNORE, LOGO_FILENAME)

    disclaimer_sheet_name = next((s for s in template_wb.sheetnames if s.strip().lower() == 'disclaimer'), None)

    if disclaimer_sheet_name:
        disclaimer_sheet = template_wb[disclaimer_sheet_name]
        disclaimer_sheet.sheet_view.showGridLines = False
        print("Hid gridlines on the Disclaimer sheet.")

    template_wb.save(output_file)
    print(f"Success! Data transferred and saved to {output_file}")
    return output_file
