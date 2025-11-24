from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

template_file = "Daily Performance Template.xlsx"
SHEETS_TO_IGNORE = ['Home', 'Sheet1', 'Disclaimer']
LOGO_FILENAME = "corpcare_logo.jpg"

# --- MAPPING & STYLING CONFIGURATION ---
RAW_TO_TEMPLATE_HEADER_MAP = {
    "Scheme Name": "Scheme Name",  "Average Maturity Years": "Average Maturity Years",
    "Modified Duration Years": "Modified Duration Years", "YTM (%)": "YTM (%)",
    "Direct Expense Ratio": "Direct Expense Ratio",
    "Latest Date": "Latest Date", "Latest NAV(Rs)": "Latest NAV(Rs)", "1 Day": "1 Day", "3 Day": "3 Day",
    "1 Week": "1 Week",
    "2 Week": "2 Week", "1 Month": "1 Month", "3 Months": "3 Months", "6 Months": "6 Months", "9 Months": "9 Months",
    "1 Year": "1 Year", "3 Years": "3 Years", "5 Years": "5 Years", "10 Years": "10 Years",
    "SINCE INCEPTION": "SINCE INCEPTION",
    "Cash & Equivalent": "Cash & Equivalent", "Others": "Others", "SOV": "SOV",
    "A1 / A1+ / A1-": "A1 / A1+ / A1-",
    "A / A+ / A1+ / A1-": "A / A+ / A1+ / A1-",
    "AA / AA+ / AA-": "AA / AA+ / AA-", "AAA": "AAA", "Unrated": "Unrated", "D": "D", "A1+ / A1-": "A1+ / A1-",
    "Exit Load": "Exit Load", "Remark": "Remark", "Inception Date": "Inception Date",
    "[Fund Manager 1]": "Fund Manager",
}
MONTH_GROUPED_HEADERS = ["Average Maturity Years", "Modified Duration Years", "YTM (%)", "Direct Expense Ratio",
                         "Cash & Equivalent", "Others", "SOV", "A / A+ / A1+ / A1-", "AA / AA+ / AA-", "AAA", "Unrated",
                         "D", "A1+ / A1-"]

RATING_HEADERS = ["A / A+ / A-", "AA / AA+ / AA-", "A / A+ / A1+ / A1-","AAA","A1 / A1+ / A1-", "Cash & Equivalent","D","Others","SOV", "Unrated"]

# --- GLOBAL STYLES ---
light_brown_fill = PatternFill(start_color="DCC7A3", end_color="DCC7A3", fill_type="solid")
no_fill = PatternFill(fill_type=None)
back_button_fill = PatternFill(start_color="DCC783", end_color="DCC783", fill_type="solid")
back_button_font = Font(name='Calibri', size=11, color="000000", bold=True, underline=None)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border_side = Side(border_style="thin", color="000000")
button_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)